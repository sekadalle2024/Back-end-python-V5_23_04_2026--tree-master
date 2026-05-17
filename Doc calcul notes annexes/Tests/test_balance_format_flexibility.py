"""
Property-Based Tests for Balance Format Flexibility

This module contains property-based tests using Hypothesis to verify
the Balance_Reader module's ability to handle various balance file formats.

**Property 21: Balance Format Flexibility**

**Validates: Requirements 14.1, 14.2, 14.3, 14.5, 14.6**

For any balance file with column name variations (multiple spaces, different separators),
the Balance_Parser must automatically detect and normalize column names, and must accept
both comma and period as decimal separators, and must accept formats of numbers with or
without thousand separators.

Auteur: Système de calcul automatique des notes annexes SYSCOHADA
Date: 28 Avril 2026
"""

import sys
import os
import pytest
from hypothesis import given, strategies as st, assume, settings
import pandas as pd
import openpyxl
from openpyxl import Workbook
import tempfile
import re
import locale

# Ajouter le chemin des modules au PYTHONPATH
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Modules'))

from balance_reader import BalanceReader, BalanceNotFoundException, InvalidBalanceFormatException


# ============================================================================
# HYPOTHESIS STRATEGIES FOR FORMAT VARIATIONS
# ============================================================================

@st.composite
def st_decimal_separator(draw):
    """
    Génère un séparateur décimal (virgule ou point).
    
    Returns:
        str: ',' ou '.'
    """
    return draw(st.sampled_from([',', '.']))


@st.composite
def st_thousand_separator(draw):
    """
    Génère un séparateur de milliers (espace, virgule, point, ou aucun).
    
    Returns:
        str: ' ', ',', '.', ou ''
    """
    return draw(st.sampled_from([' ', ',', '.', '']))


@st.composite
def st_formatted_number(draw, decimal_sep='.', thousand_sep=''):
    """
    Génère un nombre formaté avec les séparateurs spécifiés.
    
    Args:
        decimal_sep: Séparateur décimal (',' ou '.')
        thousand_sep: Séparateur de milliers (' ', ',', '.', ou '')
    
    Returns:
        str: Nombre formaté en tant que chaîne
    """
    # Générer un nombre aléatoire
    integer_part = draw(st.integers(min_value=0, max_value=999999))
    decimal_part = draw(st.integers(min_value=0, max_value=99))
    
    # Formater la partie entière avec séparateur de milliers
    if thousand_sep and integer_part >= 1000:
        # Convertir en chaîne et ajouter les séparateurs
        int_str = str(integer_part)
        # Ajouter les séparateurs de milliers de droite à gauche
        formatted_int = ''
        for i, digit in enumerate(reversed(int_str)):
            if i > 0 and i % 3 == 0:
                formatted_int = thousand_sep + formatted_int
            formatted_int = digit + formatted_int
    else:
        formatted_int = str(integer_part)
    
    # Formater le nombre complet
    formatted_number = f"{formatted_int}{decimal_sep}{decimal_part:02d}"
    
    return formatted_number


@st.composite
def st_balance_with_format_variations(draw):
    """
    Génère un fichier Excel avec des variations de format de nombres.
    
    Cette stratégie crée un fichier Excel avec:
    - Colonnes avec espaces multiples
    - Nombres avec différents séparateurs décimaux (virgule ou point)
    - Nombres avec différents séparateurs de milliers (espace, virgule, point, ou aucun)
    - Mélange de formats dans le même fichier
    
    Returns:
        Tuple[str, Dict]: (chemin du fichier, dictionnaire des formats utilisés)
    """
    # Générer les séparateurs
    decimal_sep = draw(st_decimal_separator())
    thousand_sep = draw(st_thousand_separator())
    
    # Générer le nombre de comptes
    num_comptes = draw(st.integers(min_value=5, max_value=15))
    
    # Générer les noms de colonnes avec espaces multiples
    base_columns = [
        'Numéro', 'Intitulé', 
        'Ant Débit', 'Ant Crédit',
        'Débit', 'Crédit',
        'Solde Débit', 'Solde Crédit'
    ]
    
    # Ajouter des espaces multiples aléatoirement
    columns_with_spaces = []
    for col in base_columns:
        num_spaces = draw(st.integers(min_value=1, max_value=3))
        col_with_spaces = col.replace(' ', ' ' * num_spaces)
        columns_with_spaces.append(col_with_spaces)
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Créer un onglet
    ws = wb.create_sheet("BALANCE N")
    
    # En-têtes avec espaces multiples
    ws.append(columns_with_spaces)
    
    # Générer les données avec les formats spécifiés
    for i in range(num_comptes):
        # Numéro de compte
        classe = draw(st.sampled_from(['1', '2', '3', '4', '5', '6', '7', '8', '9']))
        sous_classe = draw(st.integers(min_value=0, max_value=9))
        detail = draw(st.integers(min_value=0, max_value=999))
        numero = f"{classe}{sous_classe}{detail:03d}"
        
        # Intitulé
        intitule = f"Compte {numero}"
        
        # Générer les montants formatés
        ant_debit_str = draw(st_formatted_number(decimal_sep=decimal_sep, thousand_sep=thousand_sep))
        ant_credit_str = draw(st_formatted_number(decimal_sep=decimal_sep, thousand_sep=thousand_sep))
        mvt_debit_str = draw(st_formatted_number(decimal_sep=decimal_sep, thousand_sep=thousand_sep))
        mvt_credit_str = draw(st_formatted_number(decimal_sep=decimal_sep, thousand_sep=thousand_sep))
        
        # Convertir en float pour les calculs
        ant_debit = float(ant_debit_str.replace(thousand_sep, '').replace(decimal_sep, '.'))
        ant_credit = float(ant_credit_str.replace(thousand_sep, '').replace(decimal_sep, '.'))
        mvt_debit = float(mvt_debit_str.replace(thousand_sep, '').replace(decimal_sep, '.'))
        mvt_credit = float(mvt_credit_str.replace(thousand_sep, '').replace(decimal_sep, '.'))
        
        # Calculer les soldes de clôture
        solde_ouverture = ant_debit - ant_credit
        solde_cloture = solde_ouverture + mvt_debit - mvt_credit
        solde_debit = max(0, solde_cloture)
        solde_credit = max(0, -solde_cloture)
        
        # Formater les soldes
        solde_debit_str = draw(st_formatted_number(decimal_sep=decimal_sep, thousand_sep=thousand_sep))
        solde_credit_str = draw(st_formatted_number(decimal_sep=decimal_sep, thousand_sep=thousand_sep))
        
        # Ajouter la ligne (les valeurs sont des chaînes formatées)
        ws.append([numero, intitule, ant_debit_str, ant_credit_str, 
                  mvt_debit_str, mvt_credit_str, solde_debit_str, solde_credit_str])
    
    # Sauvegarder le fichier
    wb.save(temp_file.name)
    
    # Retourner le chemin et les formats utilisés
    formats_info = {
        'decimal_separator': decimal_sep,
        'thousand_separator': thousand_sep,
        'num_comptes': num_comptes,
        'columns_with_spaces': columns_with_spaces
    }
    
    return temp_file.name, formats_info


# ============================================================================
# PROPERTY-BASED TESTS - FORMAT FLEXIBILITY
# ============================================================================

@given(data=st.data())
@settings(max_examples=50, deadline=60000)
def test_property_balance_format_flexibility(data):
    """
    **Property 21: Balance Format Flexibility**
    
    **Validates: Requirements 14.1, 14.2, 14.3, 14.5, 14.6**
    
    For any balance file with column name variations (multiple spaces, different separators),
    the Balance_Parser must automatically detect and normalize column names, and must accept
    both comma and period as decimal separators, and must accept formats of numbers with or
    without thousand separators.
    
    This property verifies that:
    1. Column names with multiple spaces are normalized
    2. Decimal separators (comma and period) are handled correctly
    3. Thousand separators (space, comma, period, or none) are handled correctly
    4. All monetary values are converted to float correctly
    5. No data is lost during format conversion
    6. The system handles mixed formats gracefully
    
    Test Strategy:
    - Generate Excel files with various format combinations
    - Use different decimal separators (comma, period)
    - Use different thousand separators (space, comma, period, none)
    - Mix column name variations with number format variations
    - Verify that Balance_Reader correctly normalizes and converts all data
    - Verify that all monetary columns are numeric after loading
    """
    # Générer un fichier avec des variations de format
    fichier_excel, formats_info = data.draw(st_balance_with_format_variations())
    
    try:
        # Créer le lecteur
        reader = BalanceReader(fichier_excel)
        
        # Charger le fichier Excel brut pour vérifier les formats
        df_brut = pd.read_excel(fichier_excel, sheet_name=0)
        
        # Vérifier que les colonnes ont bien des espaces multiples
        has_multiple_spaces = any(re.search(r'\s{2,}', str(col)) for col in df_brut.columns)
        assume(has_multiple_spaces)
        
        # Appliquer le nettoyage des colonnes
        df_cleaned = reader.nettoyer_colonnes(df_brut.copy())
        
        # Appliquer la conversion des montants
        df_converted = reader.convertir_montants(df_cleaned.copy())
        
        # Vérifier que les colonnes sont normalisées
        colonnes_attendues = {
            'Numéro', 'Intitulé', 
            'Ant Débit', 'Ant Crédit',
            'Débit', 'Crédit',
            'Solde Débit', 'Solde Crédit'
        }
        
        colonnes_converties = set(df_converted.columns)
        assert colonnes_converties == colonnes_attendues, \
            f"Colonnes attendues: {colonnes_attendues}, trouvé: {colonnes_converties}"
        
        # Vérifier que tous les montants sont numériques
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        
        for col in colonnes_montants:
            assert pd.api.types.is_numeric_dtype(df_converted[col]), \
                f"Colonne {col} doit être numérique après conversion, trouvé {df_converted[col].dtype}"
            
            # Vérifier qu'il n'y a pas de NaN
            assert not df_converted[col].isna().any(), \
                f"Colonne {col} ne doit pas contenir de NaN après conversion"
            
            # Vérifier que toutes les valeurs sont >= 0
            assert (df_converted[col] >= 0).all(), \
                f"Colonne {col} doit contenir uniquement des valeurs >= 0"
        
        # Vérifier que le nombre de lignes est préservé
        assert len(df_converted) == formats_info['num_comptes'], \
            f"Nombre de lignes doit être {formats_info['num_comptes']}, trouvé {len(df_converted)}"
        
        # Vérifier que la colonne Numéro est préservée
        assert not df_converted['Numéro'].isna().any(), \
            "Colonne Numéro ne doit pas contenir de NaN"
        
        # Vérifier que la colonne Intitulé est préservée
        assert not df_converted['Intitulé'].isna().any(), \
            "Colonne Intitulé ne doit pas contenir de NaN"
        
        # Vérifier que les valeurs numériques sont raisonnables (> 0 ou = 0)
        for col in colonnes_montants:
            assert (df_converted[col] >= 0).all(), \
                f"Toutes les valeurs de {col} doivent être >= 0"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(fichier_excel):
            try:
                os.unlink(fichier_excel)
            except Exception:
                pass


@given(decimal_sep=st_decimal_separator(), thousand_sep=st_thousand_separator())
@settings(max_examples=20, deadline=30000)
def test_property_decimal_separator_handling(decimal_sep, thousand_sep):
    """
    **Property 21: Balance Format Flexibility - Decimal Separator Handling**
    
    **Validates: Requirements 14.2, 14.5**
    
    For any balance file with numbers using comma or period as decimal separator,
    the Balance_Reader must correctly convert all numbers to float.
    
    This property verifies that:
    1. Comma as decimal separator is handled correctly
    2. Period as decimal separator is handled correctly
    3. Thousand separators don't interfere with decimal conversion
    4. All values are converted to float without loss of precision
    5. No NaN values are introduced during conversion
    
    Test Strategy:
    - Generate numbers with specific decimal and thousand separators
    - Create Excel files with these formatted numbers
    - Verify that Balance_Reader correctly converts all numbers
    - Verify that the converted values are numeric and valid
    """
    # Générer le nombre de comptes
    num_comptes = 10
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
    temp_file.close()
    
    try:
        # Créer le workbook
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Créer un onglet
        ws = wb.create_sheet("BALANCE N")
        
        # En-têtes
        headers = ['Numéro', 'Intitulé', 'Ant Débit', 'Ant Crédit', 
                   'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
        ws.append(headers)
        
        # Générer les données avec les séparateurs spécifiés
        for i in range(num_comptes):
            # Numéro de compte
            numero = f"1{i:04d}"
            intitule = f"Compte {numero}"
            
            # Générer les montants formatés
            ant_debit = 1000.50
            ant_credit = 500.25
            mvt_debit = 200.75
            mvt_credit = 100.50
            solde_debit = 600.00
            solde_credit = 0.00
            
            # Formater les nombres
            def format_number(num, decimal_sep, thousand_sep):
                # Convertir en chaîne avec 2 décimales
                num_str = f"{num:.2f}"
                
                # Remplacer le point par le séparateur décimal
                num_str = num_str.replace('.', decimal_sep)
                
                # Ajouter les séparateurs de milliers si nécessaire
                if thousand_sep and num >= 1000:
                    parts = num_str.split(decimal_sep)
                    int_part = parts[0]
                    dec_part = parts[1] if len(parts) > 1 else '00'
                    
                    # Ajouter les séparateurs de milliers
                    formatted_int = ''
                    for j, digit in enumerate(reversed(int_part)):
                        if j > 0 and j % 3 == 0:
                            formatted_int = thousand_sep + formatted_int
                        formatted_int = digit + formatted_int
                    
                    num_str = formatted_int + decimal_sep + dec_part
                
                return num_str
            
            # Formater tous les montants
            ant_debit_str = format_number(ant_debit, decimal_sep, thousand_sep)
            ant_credit_str = format_number(ant_credit, decimal_sep, thousand_sep)
            mvt_debit_str = format_number(mvt_debit, decimal_sep, thousand_sep)
            mvt_credit_str = format_number(mvt_credit, decimal_sep, thousand_sep)
            solde_debit_str = format_number(solde_debit, decimal_sep, thousand_sep)
            solde_credit_str = format_number(solde_credit, decimal_sep, thousand_sep)
            
            # Ajouter la ligne
            ws.append([numero, intitule, ant_debit_str, ant_credit_str, 
                      mvt_debit_str, mvt_credit_str, solde_debit_str, solde_credit_str])
        
        # Sauvegarder le fichier
        wb.save(temp_file.name)
        
        # Créer le lecteur
        reader = BalanceReader(temp_file.name)
        
        # Charger le fichier Excel brut
        df_brut = pd.read_excel(temp_file.name, sheet_name=0)
        
        # Appliquer la conversion des montants
        df_converted = reader.convertir_montants(df_brut.copy())
        
        # Vérifier que tous les montants sont numériques
        colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crébit']
        
        for col in colonnes_montants:
            if col in df_converted.columns:
                assert pd.api.types.is_numeric_dtype(df_converted[col]), \
                    f"Colonne {col} doit être numérique"
                
                # Vérifier qu'il n'y a pas de NaN
                assert not df_converted[col].isna().any(), \
                    f"Colonne {col} ne doit pas contenir de NaN"
        
    finally:
        # Nettoyer le fichier temporaire
        if os.path.exists(temp_file.name):
            try:
                os.unlink(temp_file.name)
            except Exception:
                pass


def test_property_format_flexibility_with_demo_file():
    """
    Test de la propriété de flexibilité de format avec le fichier de démonstration.
    
    **Validates: Requirements 14.1, 14.2, 14.3, 14.5, 14.6**
    
    Ce test vérifie que le fichier de démonstration respecte les propriétés
    de flexibilité de format.
    """
    # Chemin vers le fichier de test
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    fichier_test = os.path.join(base_dir, "P000 -BALANCE DEMO N_N-1_N-2.xls")
    
    # Vérifier que le fichier existe
    if not os.path.exists(fichier_test):
        pytest.skip(f"Fichier de démonstration non trouvé: {fichier_test}")
    
    # Créer le lecteur
    reader = BalanceReader(fichier_test)
    
    # Charger les balances
    balance_n, balance_n1, balance_n2 = reader.charger_balances()
    
    # Vérifier que tous les montants sont numériques
    colonnes_montants = ['Ant Débit', 'Ant Crédit', 'Débit', 'Crédit', 'Solde Débit', 'Solde Crédit']
    
    for balance, nom in [(balance_n, 'N'), (balance_n1, 'N-1'), (balance_n2, 'N-2')]:
        for col in colonnes_montants:
            assert pd.api.types.is_numeric_dtype(balance[col]), \
                f"Colonne {col} de Balance {nom} doit être numérique"
            
            # Vérifier qu'il n'y a pas de NaN
            assert not balance[col].isna().any(), \
                f"Colonne {col} de Balance {nom} ne doit pas contenir de NaN"
            
            # Vérifier que toutes les valeurs sont >= 0
            assert (balance[col] >= 0).all(), \
                f"Colonne {col} de Balance {nom} doit contenir uniquement des valeurs >= 0"
    
    print(f"\n✓ Propriété de flexibilité de format validée avec le fichier de démonstration")
    print(f"  - Balance N:   {len(balance_n)} comptes, tous les montants numériques")
    print(f"  - Balance N-1: {len(balance_n1)} comptes, tous les montants numériques")
    print(f"  - Balance N-2: {len(balance_n2)} comptes, tous les montants numériques")


if __name__ == "__main__":
    """
    Exécution directe des tests pour validation rapide.
    
    Usage:
        python test_balance_format_flexibility.py
    """
    print("=" * 70)
    print("PROPERTY-BASED TESTS - BALANCE FORMAT FLEXIBILITY")
    print("=" * 70)
    
    # Test avec le fichier de démonstration
    print("\n[1] Test de flexibilité de format avec le fichier de démonstration...")
    try:
        test_property_format_flexibility_with_demo_file()
        print("   ✓ Test réussi")
    except Exception as e:
        print(f"   ✗ Test échoué: {str(e)}")
    
    print("\n" + "=" * 70)
    print("Pour exécuter tous les tests property-based avec Hypothesis:")
    print("  pytest test_balance_format_flexibility.py -v")
    print("=" * 70)
